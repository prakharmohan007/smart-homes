from openpyxl import Workbook
import glob
import re


# Class to create xlsx file with room IDs every 5 secs for 30 days
class XLSXWriter:
    def __init__(self, subject_id=2,
                 dir_name="../data/experimental_data/Subject_2/processed_data",
                 num_days=30,
                 scale = 10):
        self.subject_id = subject_id
        self.dir = dir_name
        self.num_days = num_days
        self.book = Workbook()
        self.week = 0
        self.f_no = 0
        self.scale = scale
        # sheet = book.active

    def set_parameters(self, subject_id, dir_name, num_days, scale):
        self.subject_id = subject_id
        self.dir = dir_name
        self.num_days = num_days
        self.scale = scale

    def sec_time(self, sec):
        ss = int(sec % 60)
        sec = int(sec / 60)
        mm = int(sec % 60)
        hh = int(sec / 60)
        time = str(hh)+":"+str(mm)+":"+str(ss)
        return time

    def write_time(self, sheet):
        row = 1
        num_cols = int(24*60*60 / self.scale)
        sec = 0
        for col in range(1, num_cols+1):
            sheet.cell(row=row, column=col).value = self.sec_time(sec)
            sec = sec + self.scale

    def generate_successive_sheet(self, files=None):
        book = Workbook()
        self.week = 1
        try:
            while (self.week - 1)*7+self.num_days < len(files):
                self.f_no = (self.week - 1)*7
                sheet = book.create_sheet("Week_"+str(self.week))
                print("[XLSXWriter] generate_sheet: Writing sheet ", self.week)
                self.write_time(sheet)

                row = 2
                for f in range(self.f_no, self.f_no+self.num_days):
                    f_name = files[f].split(sep='/')[-1]
                    f_name = f_name.split(sep='_')[-1]
                    f_name = f_name.split(sep='.')[0]

                    sheet.cell(row=row, column=1).value = f_name

                    # open routine file
                    with open(files[f], 'r') as f_read:
                        data = f_read.readlines()
                    del data[0]

                    col = 2
                    jump = 1
                    for line in data:

                        if jump == int(self.scale / 5):
                            jump = 0
                        else:
                            sample = re.split(',', line)
                            # print(sample[-1].splitlines())
                            sheet.cell(row=row, column=col).value = sample[-1].splitlines()[0]
                            col = col + 1
                        jump = jump + 1
                    row = row + 1
                self.week = self.week + 1
            book.save("../data/experimental_data/Subject_2/ground_truth.xlsx")
        except Exception as err:
            print("[XLSXWriter] generate_first_sheet: Error: ", err)
            raise

    def generate_first_sheet(self, files=None):
        self.week = 1
        sheet = self.book.active
        sheet.tile = str(self.week)+"-"+str(self.week+3)

        print("[XLSXWriter] generate_first_sheet: Writing first sheet")

        self.write_time(sheet)

        try:
            row = 2
            for f in range(self.f_no, self.f_no+self.num_days):
                # open routine file

                with open(files[f], 'r') as f_read:
                    data = f_read.readlines()
                del data[0]
                col = 1

                jump = 1
                for line in data:

                    if jump == int(self.scale/5):
                        jump = 0
                    else:
                        sample = re.split(',', line)
                        # print(sample[-1].splitlines())
                        sheet.cell(row=row, column=col).value = sample[-1].splitlines()[0]
                        col = col + 1
                    jump = jump + 1
                row = row + 1
        except Exception as err:
            print("[XLSXWriter] generate_first_sheet: Error: ", err)
            raise
        print("[XLSXWrite] generate_first_xlsx: Number of Columns: ", sheet.max_column)
        print("[XLSXWrite] generate_first_xlsx: Number of Rows: ", sheet.max_row)

    # Read files in to a list.
    # pass the list for 1st xlsx sheet
    def generate_xlsx(self):
        # Read Files
        print("[XLSXWriter] generate_xlsx: Reading files")
        log_files = glob.glob(self.dir + "/*.log")
        if len(log_files) == 0:
            print("[XLSXWriter] generate_xlsx: Error - No log files found in the folder ", self.dir)
            exit(-1)

        log_files.sort()
        # self.generate_first_sheet(log_files)
        self.generate_successive_sheet(log_files)
        # self.book.save("../data/experimental_data/Subject_2/ground_truth.xlsx")


if __name__ == "__main__":
    obj = XLSXWriter(subject_id=2, dir_name="../data/experimental_data/Subject_2/processed_data", num_days=30)
    obj.generate_xlsx()

    # book = Workbook()
    # sheet = book.active
    #
    # sheet['A1'] = 1
    # sheet.cell(row=2, column=2).value = 2
    # sheet.title = str(1)+'-'+str(5)
    # book.save('../data/sample.xlsx')

    exit(1)
